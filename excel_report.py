"""
excel_report.py — Colorful Excel reports for Harar NID Bot
"""

import os
from datetime import date, datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import database as db

# ── colors ────────────────────────────────────────────────────────────────────
C_HEADER   = "1A237E"   # deep indigo
C_TITLE    = "283593"
C_SUBHEAD  = "3949AB"
C_ALT1     = "E8EAF6"
C_ALT2     = "FFFFFF"
C_TOTAL    = "E65100"   # deep orange
C_ZERO     = "FFCDD2"   # light red
C_HIGH     = "C8E6C9"   # light green
C_MID      = "FFF9C4"   # light yellow
C_WHITE    = "FFFFFF"

def fill(c): return PatternFill("solid", fgColor=c)
def font(bold=True, size=11, color="000000"): return Font(bold=bold, size=size, color=color)
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():   return Alignment(horizontal="left",   vertical="center")

THIN   = Side(style="thin",   color="9FA8DA")
MEDIUM = Side(style="medium", color="3949AB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MBORDER= Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)

def hrow(ws, row, values, bg=C_HEADER, fg=C_WHITE, size=11, height=22):
    ws.row_dimensions[row].height = height
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill    = fill(bg)
        c.font    = font(True, size, fg)
        c.alignment = center()
        c.border  = BORDER

def drow(ws, row, values, alt=False):
    bg = C_ALT1 if alt else C_ALT2
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        cell_bg = bg
        if isinstance(val, int):
            if col > 3 and val == 0:   cell_bg = C_ZERO
            elif col > 3 and val >= 25: cell_bg = C_HIGH
            elif col > 3 and val >= 15: cell_bg = C_MID
        c.fill      = fill(cell_bg)
        c.font      = font(False, 11)
        c.alignment = left() if col == 2 else center()
        c.border    = BORDER

def trow(ws, row, values):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill      = fill(C_TOTAL)
        c.font      = font(True, 12, C_WHITE)
        c.alignment = center()
        c.border    = MBORDER

def widths(ws, w_list):
    for i, w in enumerate(w_list, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title_row(ws, row, text, cols, bg=C_HEADER, size=14, height=30):
    ws.merge_cells(f"A{row}:{get_column_letter(cols)}{row}")
    c = ws[f"A{row}"]
    c.value     = text
    c.fill      = fill(bg)
    c.font      = font(True, size, C_WHITE)
    c.alignment = center()
    ws.row_dimensions[row].height = height

# ══════════════════════════════════════════════════════════════════════════════

def write_day_sheet(ws, officers, date_iso):
    rpts = db.get_reports_for_date(date_iso)
    d    = datetime.strptime(date_iso, "%Y-%m-%d")
    dfmt = d.strftime("%d/%m/%Y")
    wday = d.strftime("%A")

    title_row(ws, 1, "📊  STATION 2 — Harar Post Office", 6)
    title_row(ws, 2, f"📅  {dfmt}   ({wday})", 6, bg=C_TITLE, size=12, height=22)
    hrow(ws, 3, ["No", "Name", "KIT Number", "Registered", "Uploaded", "Difference"], bg=C_SUBHEAD, size=10)

    total_r = total_u = 0
    r = 4; n = 1
    for o in officers:
        for kit in o["kits"]:
            e   = rpts.get(kit, {})
            reg = e.get("reg", 0)
            upl = e.get("uploaded", 0)
            diff= reg - upl
            total_r += reg; total_u += upl
            drow(ws, r, [n, o["name"], kit, reg, upl, diff], alt=(n % 2 == 0))
            r += 1; n += 1

    trow(ws, r, ["", "🔢  TOTAL", "", total_r, total_u, total_r - total_u])
    widths(ws, [5, 15, 13, 13, 13, 13])


def generate_daily_excel(date_iso: str) -> str:
    officers = db.get_all_officers()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Report"
    write_day_sheet(ws, officers, date_iso)
    path = f"/tmp/daily_{date_iso}.xlsx"
    wb.save(path)
    return path


def generate_weekly_excel() -> str:
    today = date.today()
    start = today - timedelta(days=today.weekday())
    days  = [(start + timedelta(days=i)).isoformat()
             for i in range(7) if (start + timedelta(days=i)) <= today]
    officers = db.get_all_officers()

    wb     = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "Weekly Summary"

    day_labels = [datetime.strptime(d, "%Y-%m-%d").strftime("%a %d/%m") for d in days]
    title_row(ws_sum, 1,
        f"📆  Weekly Report — {start.strftime('%d/%m/%Y')} to {today.strftime('%d/%m/%Y')}",
        3 + len(days) + 2)
    hrow(ws_sum, 2,
        ["No", "Name", "KIT"] + day_labels + ["Total Reg", "Total Up"],
        bg=C_SUBHEAD, size=10)

    r = 3; n = 1
    for o in officers:
        for kit in o["kits"]:
            vals = [n, o["name"], kit]
            tr = tu = 0
            for d_iso in days:
                rpts = db.get_reports_for_date(d_iso)
                e = rpts.get(kit, {})
                reg = e.get("reg", 0); upl = e.get("uploaded", 0)
                vals.append(f"{reg}/{upl}")
                tr += reg; tu += upl
            vals += [tr, tu]
            drow(ws_sum, r, vals, alt=(n % 2 == 0))
            r += 1; n += 1

    widths(ws_sum, [5, 15, 13] + [13] * len(days) + [12, 12])

    for d_iso in days:
        label = datetime.strptime(d_iso, "%Y-%m-%d").strftime("%a_%d%m")
        ws = wb.create_sheet(title=label)
        write_day_sheet(ws, officers, d_iso)

    path = f"/tmp/weekly_{start.isoformat()}.xlsx"
    wb.save(path)
    return path


def generate_monthly_excel() -> str:
    today = date.today()
    first = today.replace(day=1)
    days  = []
    d = first
    while d <= today:
        days.append(d.isoformat()); d += timedelta(days=1)

    officers = db.get_all_officers()
    wb     = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "Monthly Summary"

    title_row(ws_sum, 1, f"🗓  Monthly Report — {first.strftime('%B %Y')}", 9)
    hrow(ws_sum, 2,
        ["No","Name","KIT","Total Reg","Total Uploaded","Difference","Avg Reg/Day","Avg Up/Day","Work Days"],
        bg=C_SUBHEAD, size=10)

    r = 3; n = 1
    for o in officers:
        for kit in o["kits"]:
            tr = tu = wd = 0
            for d_iso in days:
                rpts = db.get_reports_for_date(d_iso)
                e = rpts.get(kit, {})
                reg = e.get("reg", 0); upl = e.get("uploaded", 0)
                if reg or upl: wd += 1
                tr += reg; tu += upl
            avg_r = round(tr / max(wd, 1), 1)
            avg_u = round(tu / max(wd, 1), 1)
            drow(ws_sum, r, [n, o["name"], kit, tr, tu, tr-tu, avg_r, avg_u, wd], alt=(n%2==0))
            r += 1; n += 1

    trow(ws_sum, r, ["","🔢 TOTAL","",
        sum(db.get_reports_for_date(d).get(k,{}).get("reg",0) for d in days for o in officers for k in o["kits"]),
        sum(db.get_reports_for_date(d).get(k,{}).get("uploaded",0) for d in days for o in officers for k in o["kits"]),
        "","","",""])
    widths(ws_sum, [5,15,13,13,13,13,12,12,11])

    # weekly sheets inside monthly
    ws_date = first
    while ws_date <= today:
        label = f"W{ws_date.strftime('%d%m')}"
        ws = wb.create_sheet(title=label)
        write_day_sheet(ws, officers, ws_date.isoformat())
        ws_date += timedelta(days=7)

    path = f"/tmp/monthly_{first.strftime('%Y_%m')}.xlsx"
    wb.save(path)
    return path
